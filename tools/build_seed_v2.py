"""
임대현황.xlsx → 11_seed_v2.sql 변환.

엑셀 컬럼:
  순번 / 회사명 / 임대기기 / 임대금액 / 사업자금액 / 카운터점검일 / 청구일 /
  계산서 / 입금일(익월) / 추가장당 / 주소 / 임대날짜 / 제품구성 / 선후불/보증금 / 비고

다중 라인 셀(회사명·임대기기·임대날짜) → 다수 item 으로 분해.
출력: rental_customers / rental_items / rental_assignments 시드.
"""
import re, sys, json
from pathlib import Path
from openpyxl import load_workbook

ROOT  = Path(__file__).resolve().parents[1]
XLSX  = ROOT / 'data' / '임대현황.xlsx'
OUT   = ROOT / 'tools' / 'sql' / '11_seed_v2.sql'
PEEK  = ROOT / 'tools' / '_seed_peek.json'

HEADER_ROW = 6   # "순번/회사명/..." 줄
DATA_START = 7   # 1부터 시작 (openpyxl iter_rows 는 1-indexed)

def norm(s):
    return re.sub(r'\s+', ' ', str(s or '').strip())

def split_lines(s):
    if s is None: return []
    return [x.strip() for x in str(s).replace('\r', '').split('\n') if x.strip()]

def is_meta_line(s):
    """ '(본점:1/지점:1)', '(2025년부터)' 같은 괄호 메타데이터 식별 """
    if not s: return False
    s = s.strip()
    if s.startswith('(') and s.endswith(')'):
        return True
    if re.search(r'(본점|지점|계약|연장|만기)\s*[:：]', s):
        return True
    return False

def parse_date(s):
    """ '2018.12.13' / '14.06.28' / '2026-05-01' → YYYY-MM-DD or None """
    if not s: return None
    s = str(s).strip()
    m = re.search(r'(\d{2,4})[\.\-/](\d{1,2})[\.\-/](\d{1,2})', s)
    if not m: return None
    y, mo, d = m.groups()
    y = int(y)
    if y < 100: y += 2000
    try:
        return f"{y:04d}-{int(mo):02d}-{int(d):02d}"
    except: return None

def parse_amount(s):
    if s is None: return 0
    if isinstance(s, (int, float)): return int(s)
    digits = re.sub(r'[^\d]', '', str(s))
    return int(digits) if digits else 0

def parse_deposit(s):
    """ '선불 / 20만' → 200000 / '후불 / 없음' → 0 """
    if not s: return 0, '선불'
    s = str(s)
    pay_type = '후불' if '후불' in s else '선불'
    # 보증금 부분
    m = re.search(r'/\s*(.+)$', s)
    if not m: return 0, pay_type
    rest = m.group(1).strip()
    if '없' in rest or '0' == rest: return 0, pay_type
    # '20만', '12만', '20만원' → 200000
    mm = re.search(r'(\d+(?:\.\d+)?)\s*만', rest)
    if mm: return int(float(mm.group(1)) * 10000), pay_type
    n = parse_amount(rest)
    return n, pay_type

def classify_device(name):
    """ 모델·제조사 텍스트로 카테고리/서브타입/브랜드 추정 """
    n = norm(name)
    low = n.lower()
    # 브랜드
    brand = None
    for b in ['교세라','kyocera','브라더','brother','hp','캐논','canon','삼성','samsung','신도리코','sindoh','엡손','epson','후지제록스','xerox','리코','ricoh','오키','oki']:
        if b in low or b in n: brand = b; break
    # 카테고리 결정
    cat, sub = '출력', '복합기'
    if any(k in n for k in ['PC','컴퓨터','데스크탑','노트북','laptop','desktop']):
        cat, sub = 'IT', 'PC'
    elif any(k in n for k in ['모니터','monitor','LED','LCD']):
        cat, sub = 'IT', 'monitor'
    elif any(k in n for k in ['NAS','nas']):
        cat, sub = 'IT', 'NAS'
    elif any(k in n for k in ['웰리스','웰니스','wellness','정수','공기청정','제균']):
        cat, sub = '위생', '웰리스'
    elif any(k in n for k in ['잉크젯','잉크','J2740','T220','T310','J37','T8','J3','J57','J27','T3','T2','T8710','8710','5910','3530']):
        cat, sub = '출력', '잉크젯'
    elif any(k in n for k in ['레이저','HL-','MFC-L','laser']):
        cat, sub = '출력', '레이저'
    return cat, sub, brand

def esc(s):
    if s is None: return 'NULL'
    if isinstance(s, (int, float)): return str(int(s))
    return "'" + str(s).replace("'", "''") + "'"

def main():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    customers = []
    items = []
    assignments = []
    cid_seq = 0; iid_seq = 0; aid_seq = 0
    skipped = []

    for r_idx in range(DATA_START - 1, len(rows)):
        r = rows[r_idx]
        if not r or r[0] is None: continue
        try: seq = int(r[0])
        except: continue

        company_raw = r[1]
        device_raw  = r[2]
        amount_raw  = r[3]
        invoice_day = r[6]
        pricing     = r[9]    # 추가장당
        address     = r[10]
        install_raw = r[11]
        compo       = r[12]   # 제품구성 (요약)
        deposit_raw = r[13]
        notes_raw   = r[14]

        company_lines_all = split_lines(company_raw)
        device_lines  = split_lines(device_raw)
        install_lines = split_lines(install_raw)

        # 메타 라인 분리 — 회사명에 (본점:1/지점:1) 같은 첨부 텍스트는 메모로
        company_lines = [c for c in company_lines_all if not is_meta_line(c)]
        meta_extras   = [c for c in company_lines_all if     is_meta_line(c)]

        if not company_lines:
            skipped.append({'row': r_idx+1, 'reason': 'no company'})
            continue

        # 다중 본·지점 (회사명 라인 수 > 1) → 별도 거래처 N개
        # 단일 회사 + 다중 기기 → 한 거래처에 item N개
        if len(company_lines) == 1 and len(device_lines) >= 1:
            customer_buckets = [(company_lines[0], device_lines, install_lines)]
        else:
            # 회사명 라인 ↔ 기기 라인 1:1 매핑 시도
            customer_buckets = []
            n = max(len(company_lines), len(device_lines))
            for i in range(n):
                comp = company_lines[i] if i < len(company_lines) else company_lines[-1]
                dev  = [device_lines[i]] if i < len(device_lines) else []
                ins  = [install_lines[i]] if i < len(install_lines) else (install_lines[-1:] if install_lines else [])
                customer_buckets.append((comp, dev, ins))

        deposit_val, pay_type = parse_deposit(deposit_raw)

        for (comp_name, dev_list, ins_list) in customer_buckets:
            cid_seq += 1
            cid = f"c_{cid_seq:04d}"
            notes_parts = []
            if notes_raw: notes_parts.append(norm(notes_raw))
            if meta_extras and customer_buckets.index((comp_name, dev_list, ins_list)) == 0:
                notes_parts.append(' / '.join(meta_extras))
            customers.append({
                'id': cid,
                'company': comp_name,
                'address': norm(address),
                'invoice_day': norm(invoice_day) if invoice_day else None,
                'payment_type': pay_type,
                'deposit': deposit_val,
                'notes': ' | '.join(notes_parts) if notes_parts else None,
            })

            # 거래처당 임대료 — 단순화: 회사 단위 amount_raw 가 통합 금액이므로
            # 첫 item 에만 monthly_fee, 나머지는 0
            first_item = True
            for di, dev_name in enumerate(dev_list):
                iid_seq += 1
                iid = f"it_{iid_seq:04d}"
                cat, sub, brand = classify_device(dev_name)
                install_date = parse_date(ins_list[di] if di < len(ins_list) else (ins_list[0] if ins_list else None))
                items.append({
                    'id': iid,
                    'category': cat,
                    'subtype': sub,
                    'brand': brand,
                    'model': dev_name,
                    'install_date': install_date,
                })
                aid_seq += 1
                aid = f"a_{aid_seq:04d}"
                assignments.append({
                    'id': aid,
                    'item_id': iid,
                    'customer_id': cid,
                    'start_date': install_date,
                    'monthly_fee': parse_amount(amount_raw) if first_item else 0,
                    'notes': norm(pricing) if pricing and first_item else None,
                })
                first_item = False

    PEEK.write_text(json.dumps({
        'customers': len(customers),
        'items': len(items),
        'assignments': len(assignments),
        'skipped': skipped,
        'sample_customer': customers[:3],
        'sample_item': items[:3],
        'sample_assignment': assignments[:3],
    }, ensure_ascii=False, indent=2), encoding='utf-8')

    lines = [
        '-- ============================================================',
        '-- 11_seed_v2.sql  (auto-generated from data/임대현황.xlsx)',
        f'-- customers={len(customers)} items={len(items)} assignments={len(assignments)}',
        '-- ============================================================',
        '',
        '-- 기존 데이터 비우기 (안전: rental_customers 만 CASCADE → 나머지 모두 cascade)',
        'TRUNCATE rental_billings, rental_supplies, rental_counters, rental_assignments, rental_items, rental_customers RESTART IDENTITY CASCADE;',
        '',
        '-- ===== rental_customers =====',
    ]
    for c in customers:
        lines.append(
            f"INSERT INTO rental_customers (id, company, address, invoice_day, payment_type, deposit, notes) VALUES ("
            f"{esc(c['id'])}, {esc(c['company'])}, {esc(c['address'])}, {esc(c['invoice_day'])}, "
            f"{esc(c['payment_type'])}, {c['deposit']}, {esc(c['notes'])});"
        )
    lines += ['', '-- ===== rental_items =====']
    for it in items:
        lines.append(
            f"INSERT INTO rental_items (id, category, subtype, brand, model, install_date) VALUES ("
            f"{esc(it['id'])}, {esc(it['category'])}, {esc(it['subtype'])}, "
            f"{esc(it['brand'])}, {esc(it['model'])}, {esc(it['install_date'])});"
        )
    lines += ['', '-- ===== rental_assignments =====']
    for a in assignments:
        lines.append(
            f"INSERT INTO rental_assignments (id, item_id, customer_id, start_date, monthly_fee, notes) VALUES ("
            f"{esc(a['id'])}, {esc(a['item_id'])}, {esc(a['customer_id'])}, "
            f"{esc(a['start_date'])}, {a['monthly_fee']}, {esc(a['notes'])});"
        )
    lines += [
        '',
        '-- 확인',
        "SELECT 'customers' AS t, COUNT(*) FROM rental_customers",
        "UNION ALL SELECT 'items', COUNT(*) FROM rental_items",
        "UNION ALL SELECT 'assignments', COUNT(*) FROM rental_assignments;",
        ''
    ]
    OUT.write_text('\n'.join(lines), encoding='utf-8')
    print(f"wrote {OUT.name}  customers={len(customers)}  items={len(items)}  assignments={len(assignments)}")
    if skipped: print(f"skipped {len(skipped)} rows")

if __name__ == '__main__':
    main()
