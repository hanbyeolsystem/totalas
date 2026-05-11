"""임대현황.xlsx 구조 파악 — JSON 파일로 출력 (콘솔 인코딩 우회)."""
import json
from pathlib import Path
import openpyxl

XL = Path(r"C:\Users\UserK\Desktop\클로드코드공부\임대관리\data\임대현황.xlsx")
OUT = Path(r"C:\Users\UserK\Desktop\클로드코드공부\임대관리\tools\_peek.json")

wb = openpyxl.load_workbook(XL, data_only=True, read_only=True)

out = {"file": XL.name, "sheets": {}}
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    out["sheets"][name] = {
        "rows": len(rows),
        "sample_head": [list(r) for r in rows[:10]],
        "sample_tail": [list(r) for r in rows[-5:]] if len(rows) > 10 else [],
    }

OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
print(f"wrote {OUT}")
