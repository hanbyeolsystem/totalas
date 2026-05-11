"""
NAS 세부현황 폴더의 248개 거래처별 xlsx → 계약서 데이터 JSON 추출.

각 파일의 '계약서' 시트 구조:
  R2C3:   회사명 (헤더)
  R5C5:   회사상호      | R5C11:  주민등록번호 / 사업자번호
  R6C5:   요청자성함    | R6C11:  휴대폰 번호
  R7C5:   설치주소      | R7C11:  전화 / 팩스
  R8C5:   발행구분      | R8C11:  이메일주소
  R11~R14 (4행): 계약내용 표 (#, 렌탈물품, 기본흑/컬, 추가흑/컬, 수량, 설치비, 월렌탈료)
  R15C5:  보증금        | R15C13: 합계금액
  R33C1:  계약일 텍스트 ("2020 년 1 월 18 일")
  R117:   특약사항 1
  R118:   특약사항 2
"""
import openpyxl, json, re
from pathlib import Path

SRC = Path(r'\\192.168.0.251\업무공용\0 한별임대현황20260511\세부현황')
OUT = Path(r'C:\Users\UserK\Desktop\클로드코드공부\임대관리\tools\seed_contracts.json')

def clean(v):
    if v is None: return ''
    s = str(v).strip()
    return s

def to_int(v):
    if v in (None, '', 0): return 0
    try: return int(round(float(str(v).replace(',', '').strip())))
    except: return 0

def parse_qty(v):
    """'1 대' → 1"""
    if not v: return 0
    m = re.search(r'(\d+)', str(v))
    return int(m.group(1)) if m else 0

def parse_date_text(v):
    """'2020 년 1 월 18 일' → '2020-01-18' (실패 시 원문)"""
    if not v: return ''
    s = str(v)
    m = re.search(r'(\d{4})\s*년\s*(\d{1,2})?\s*월?\s*(\d{1,2})?', s)
    if m:
        y = m.group(1)
        mo = m.group(2) or '01'
        d = m.group(3) or '01'
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return s.strip()

def extract_contract(xlsx_path):
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    except Exception as e:
        return None, f"열기 실패: {e}"
    if '계약서' not in wb.sheetnames:
        wb.close()
        return None, "계약서 시트 없음"
    ws = wb['계약서']

    # 실제 셀 위치: 라벨이 C10, 값이 C12 (C11은 비어있음)
    def right_val(r):
        v = ws.cell(r, 12).value
        if v is None or not str(v).strip():
            v = ws.cell(r, 11).value
        return clean(v)

    company_top = clean(ws.cell(2, 3).value) or clean(ws.cell(2, 2).value)
    company     = clean(ws.cell(5, 5).value) or company_top
    biz_no      = right_val(5)
    requester   = clean(ws.cell(6, 5).value)
    mobile      = right_val(6)
    address     = clean(ws.cell(7, 5).value)
    tel_fax     = right_val(7)
    invoice_kind= clean(ws.cell(8, 5).value)
    email       = right_val(8)

    items = []
    for r in (11, 12, 13, 14):
        no = ws.cell(r, 2).value
        product = clean(ws.cell(r, 3).value)
        if not product and not ws.cell(r, 13).value:
            continue
        items.append({
            'no':       to_int(no) or (r - 10),
            'product':  product,
            'bw_free':  to_int(ws.cell(r, 6).value),
            'co_free':  to_int(ws.cell(r, 7).value),
            'bw_rate':  to_int(ws.cell(r, 8).value),
            'co_rate':  to_int(ws.cell(r, 9).value),
            'qty':      parse_qty(ws.cell(r, 10).value),
            'install':  clean(ws.cell(r, 12).value) or '무료',
            'fee':      to_int(ws.cell(r, 13).value),
            'vat_note': clean(ws.cell(r, 14).value),
        })

    deposit       = to_int(ws.cell(15, 5).value)
    total_fee     = to_int(ws.cell(15, 13).value)
    contract_date = parse_date_text(ws.cell(33, 1).value)
    spec1         = clean(ws.cell(117, 1).value)
    spec2         = clean(ws.cell(118, 1).value)
    special = []
    if spec1 and spec1 != '1': special.append(spec1)
    if spec2 and spec2 not in ('2',) and '이를 증명' not in spec2: special.append(spec2)

    wb.close()
    return {
        'company': company,
        'company_top': company_top,
        'requester': requester,
        'address': address,
        'invoice_kind': invoice_kind,
        'biz_no': biz_no,
        'mobile': mobile,
        'tel_fax': tel_fax,
        'email': email,
        'items': items,
        'deposit': deposit,
        'total_fee': total_fee,
        'contract_date': contract_date,
        'special': special,
        'source_file': xlsx_path.name,
    }, None

def main():
    if not SRC.exists():
        print(f"NAS 경로 접근 불가: {SRC}")
        return
    files = [p for p in sorted(SRC.glob('*.xlsx')) if not p.name.startswith('~$')]
    print(f"발견: {len(files)}개 파일")

    contracts = []
    errors = []
    for i, p in enumerate(files):
        if i % 20 == 0: print(f"  ... {i}/{len(files)}")
        c, err = extract_contract(p)
        if err:
            errors.append({'file': p.name, 'error': err})
            continue
        if c: contracts.append(c)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({
        'extracted_at': '2026-05-11',
        'source_dir': str(SRC),
        'count': len(contracts),
        'errors': errors[:30],
        'rows': contracts,
    }, ensure_ascii=False, indent=2), encoding='utf-8')

    print(f"\n=== 계약서 추출 완료 ===")
    print(f"  성공:  {len(contracts)}건")
    print(f"  오류:  {len(errors)}건")
    print(f"  출력:  {OUT}")
    if errors[:5]:
        print(f"\n  오류 샘플:")
        for e in errors[:5]: print(f"    {e['file']}: {e['error']}")

    # 샘플 5건
    print("\n=== 샘플 ===")
    for c in contracts[:5]:
        items_str = ', '.join(f"{it['product']}({it['fee']})" for it in c['items'] if it['product'])
        print(f"  {c['company']:25} | {c['contract_date']:12} | 합계 {c['total_fee']:>8} | {items_str}")

if __name__ == '__main__':
    main()
