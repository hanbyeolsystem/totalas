"""거래처별 추가카운트 엑셀 샘플 3개 — 구조 파악."""
import json
from pathlib import Path
import openpyxl

D = Path(r"C:\Users\UserK\Desktop\클로드코드공부\임대관리\data\거래처별 추가카운트")
OUT = Path(r"C:\Users\UserK\Desktop\클로드코드공부\임대관리\tools\_extra_peek.json")

samples = [
    "추가카운터_주은교육.xlsx",
    "추가카운터_그린노인요양원.xlsx",
    "추가카운터_극동볼트.xlsx",
    "추가카운터_콘텐츠랩 슬러그.xlsx",
]
result = {}
for name in samples:
    p = D / name
    if not p.exists():
        result[name] = {"error": "not found"}
        continue
    try:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    except Exception as e:
        result[name] = {"error": str(e)}
        continue
    sheets = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        sheets[sn] = {
            "rows": len(rows),
            "head": [list(r) for r in rows[:25]],
        }
    result[name] = sheets

OUT.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str),
                encoding="utf-8")
print(f"wrote {OUT.name}")
