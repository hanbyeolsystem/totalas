"""
월별카운터/*/한별카운터_*.xlsx → 15_seed_counters.sql

매칭 키:
  rental_items.model 의 모델넘버 + 시드된 회사명 (rental_customers.company)
  ↔ 한별카운터 엑셀의 모델/자산번호.

출력:
  tools/sql/15_seed_counters.sql   — rental_items.serial UPDATE + rental_counters UPSERT
  tools/_counter_match_report.json — 매칭/미매칭 리포트
"""
import re, json
from pathlib import Path
from difflib import SequenceMatcher
from openpyxl import load_workbook

ROOT  = Path(__file__).resolve().parents[1]
SEED  = ROOT / 'tools' / 'sql' / '11_seed_v2.sql'
MONTH_DIR = ROOT / 'data' / '월별카운터'
OUT_SQL   = ROOT / 'tools' / 'sql' / '15_seed_counters.sql'
OUT_RPT   = ROOT / 'tools' / '_counter_match_report.json'
MAP_FILE  = ROOT / 'tools' / 'manual_company_map.json'  # 선택: 자산번호 → 회사명 명시 매핑

# ---------- 1. 11_seed_v2.sql 파싱 ----------
def parse_seed():
    text = SEED.read_text(encoding='utf-8')
    # rental_customers: (id, company, address, invoice_day, payment_type, deposit, notes)
    cust_re = re.compile(
        r"INSERT INTO rental_customers \([^)]+\) VALUES \('(?P<id>c_\d+)', '(?P<company>(?:[^']|'')*)',",
    )
    customers = {m.group('id'): m.group('company').replace("''", "'") for m in cust_re.finditer(text)}

    # rental_items: (id, category, subtype, brand, model, install_date, status)
    item_re = re.compile(
        r"INSERT INTO rental_items \(([^)]+)\) VALUES \(([^;]+)\);"
    )
    items = {}     # iid -> {model, brand, install_date}
    for m in item_re.finditer(text):
        cols = [c.strip() for c in m.group(1).split(',')]
        # 값 파싱은 단순화 — '...', ... 형태에서 따옴표 처리
        vals = _split_sql_values(m.group(2))
        rec = dict(zip(cols, vals))
        items[rec['id']] = {
            'model': rec.get('model'),
            'brand': rec.get('brand'),
            'install_date': rec.get('install_date'),
        }

    # rental_assignments: customer_id <-> item_id
    asgn_re = re.compile(
        r"INSERT INTO rental_assignments \(([^)]+)\) VALUES \(([^;]+)\);"
    )
    item_to_cust = {}
    for m in asgn_re.finditer(text):
        cols = [c.strip() for c in m.group(1).split(',')]
        vals = _split_sql_values(m.group(2))
        rec = dict(zip(cols, vals))
        item_to_cust[rec['item_id']] = rec['customer_id']

    out = []
    for iid, it in items.items():
        cid = item_to_cust.get(iid)
        comp = customers.get(cid) if cid else None
        out.append({
            'item_id': iid,
            'customer_id': cid,
            'company': comp,
            'model': it['model'],
            'brand': it['brand'],
        })
    return out

def _split_sql_values(s):
    """SQL VALUES 안의 콤마 split — 따옴표 보호."""
    out = []
    buf = ''
    in_str = False
    i = 0
    while i < len(s):
        ch = s[i]
        if in_str:
            if ch == "'" and i+1 < len(s) and s[i+1] == "'":
                buf += ch; i += 2; continue
            if ch == "'":
                in_str = False
                buf += ch
            else:
                buf += ch
            i += 1; continue
        if ch == "'":
            in_str = True
            buf += ch
            i += 1; continue
        if ch == ',':
            out.append(buf.strip())
            buf = ''
            i += 1; continue
        buf += ch
        i += 1
    if buf.strip(): out.append(buf.strip())
    # 값 디코딩
    decoded = []
    for v in out:
        if v == 'NULL': decoded.append(None)
        elif v.startswith("'") and v.endswith("'"):
            decoded.append(v[1:-1].replace("''", "'"))
        else:
            try: decoded.append(int(v))
            except: decoded.append(v)
    return decoded

# ---------- 2. 한별카운터 엑셀 파싱 ----------
def parse_counter_xlsx(path):
    """ Row: 그룹 / 그룹경로 / 모델 / 시리얼 / (none) / 자산번호 / IP / (none) / 위치 / 마지막업데이트 / (none) / 획득일 / 결합합계 / 흑백합계 / 컬러합계 / ... """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    HEADER_ROW = 8  # 1-based
    rows = []
    for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if not r or not r[2]:  # 모델 비어있으면 skip
            continue
        if r[0] and str(r[0]).startswith('소스'): break  # 합계 행
        rows.append({
            'group':  r[0],
            'model':  str(r[2]).strip(),
            'serial': str(r[3]).strip() if r[3] else None,
            'asset':  str(r[5]).strip() if r[5] else None,  # 자산 번호 = 거래처/위치명
            'ip':     r[6],
            'last_update': r[9],
            'bw':    _to_int(r[13]),
            'color': _to_int(r[14]),
        })
    return rows

def _to_int(v):
    if v is None or v == '': return 0
    try: return int(v)
    except:
        try: return int(float(v))
        except: return 0

# ---------- 3. 매칭 ----------
def normalize(s):
    if not s: return ''
    s = str(s).lower()
    s = re.sub(r'[\s\(\)\[\]\-_.,/:]+', '', s)
    return s

# 모델 번호만 추출 (3554ci, 2554ci, J2740, T220 등 영문/숫자 토큰)
MODEL_TOKEN = re.compile(r'[a-zA-Z]*\d+[a-zA-Z]+\d*|\d{3,5}[a-zA-Z]*')

def model_tokens(s):
    if not s: return set()
    return {normalize(t) for t in MODEL_TOKEN.findall(str(s))}

def sim(a, b):
    return SequenceMatcher(None, normalize(a), normalize(b)).ratio()

def load_manual_map():
    if not MAP_FILE.exists(): return {}
    raw = json.loads(MAP_FILE.read_text(encoding='utf-8'))
    # _ 로 시작하는 키(주석/메모)는 제거
    return {k: v for k, v in raw.items() if not k.startswith('_') and isinstance(v, str)}

def match_one(counter_row, items, manual_map):
    asset = counter_row.get('asset') or ''
    cn_company = normalize(asset)
    cn_model_tokens = model_tokens(counter_row.get('model'))

    # 0) 매뉴얼 매핑 우선 (자산번호 → 회사명 강제 지정)
    forced_company = manual_map.get(asset) or manual_map.get(asset.strip())
    if forced_company:
        fc = normalize(forced_company)
        candidates = [it for it in items if normalize(it.get('company')) == fc]
        if candidates:
            if len(candidates) == 1: return candidates[0], 'manual'
            # 모델 토큰으로 disambiguate
            scored = sorted(
                ((len(cn_model_tokens & model_tokens(it.get('model'))), it) for it in candidates),
                key=lambda x: -x[0])
            if scored[0][0] > 0: return scored[0][1], 'manual+model'
            return candidates[0], 'manual-first'

    # 1) 정확/포함 매칭
    exact, partial = [], []
    for it in items:
        ic = normalize(it.get('company'))
        if not ic or not cn_company: continue
        if ic == cn_company:        exact.append(it)
        elif ic in cn_company or cn_company in ic: partial.append(it)
    candidates = exact or partial

    # 2) 유사도 매칭 (둘 다 비었을 때) — 임계 0.62
    if not candidates:
        scored = sorted(
            ((sim(it.get('company'), asset), it) for it in items),
            key=lambda x: -x[0])
        if scored and scored[0][0] >= 0.62:
            top = scored[0][0]
            close = [it for s, it in scored if s >= top - 0.05]
            candidates = close

    if not candidates: return None, 'no-company'
    if len(candidates) == 1: return candidates[0], 'company-only'

    # 3) 모델 토큰으로 disambiguate
    scored = sorted(
        ((len(cn_model_tokens & model_tokens(it.get('model'))), it) for it in candidates),
        key=lambda x: -x[0])
    if scored[0][0] > 0: return scored[0][1], 'model-token'

    # 4) 마지막 fallback — 회사명 유사도 가장 높은 1개
    scored = sorted(
        ((sim(it.get('company'), asset), it) for it in candidates),
        key=lambda x: -x[0])
    if scored and scored[0][0] >= 0.62:
        return scored[0][1], 'sim-best'
    return None, 'ambiguous'

# ---------- 4. SQL escape ----------
def esc(v):
    if v is None: return 'NULL'
    if isinstance(v, (int, float)): return str(int(v))
    return "'" + str(v).replace("'", "''") + "'"

# ---------- 5. 메인 ----------
def main():
    items = parse_seed()
    manual = load_manual_map()
    print(f'시드된 items: {len(items)}  / 매뉴얼 매핑: {len(manual)}')

    # 월별 폴더 → ym 매핑
    folder_to_ym = {
        '2026_01': '2026-01',
        '2026_02': '2026-02',
        '2026_03': '2026-03',
        '2026_04': '2026-04',
    }
    files = []
    for folder, ym in folder_to_ym.items():
        d = MONTH_DIR / folder
        if not d.exists(): continue
        for f in d.glob('한별카운터_*.xlsx'):
            files.append((ym, f))

    report = {'months': {}, 'serial_updates': {}, 'counters': []}
    sql_lines = []
    sql_lines.append('-- ============================================================')
    sql_lines.append('-- 15_seed_counters.sql  (auto-generated)')
    sql_lines.append('-- data/월별카운터/*/한별카운터_*.xlsx → rental_counters + serial UPDATE')
    sql_lines.append('-- 매칭: rental_customers.company × rental_items.model 토큰')
    sql_lines.append('-- ============================================================')
    sql_lines.append('')

    item_to_serial = {}  # iid -> first-seen serial

    for ym, f in sorted(files):
        rows = parse_counter_xlsx(f)
        matched, missed = [], []
        for r in rows:
            it, mode = match_one(r, items, manual)
            if it:
                matched.append({'item_id': it['item_id'], 'company': it['company'], 'model_it': it['model'],
                               'model_xlsx': r['model'], 'serial': r['serial'], 'bw': r['bw'], 'color': r['color'],
                               'asset': r['asset'], 'mode': mode})
                # 시리얼 보강 (첫 발견 우선)
                if r['serial'] and it['item_id'] not in item_to_serial:
                    item_to_serial[it['item_id']] = r['serial']
            else:
                missed.append({'asset': r['asset'], 'model': r['model'], 'serial': r['serial'],
                              'bw': r['bw'], 'color': r['color'], 'reason': mode})
        report['months'][ym] = {
            'file': f.name, 'total': len(rows),
            'matched': len(matched), 'missed': len(missed),
            'missed_samples': missed[:20],
        }
        report['counters'].extend([{'ym': ym, **m} for m in matched])

    # serial UPDATE
    sql_lines.append('-- ---- rental_items.serial 보강 ----')
    for iid, sn in item_to_serial.items():
        sql_lines.append(f"UPDATE rental_items SET serial={esc(sn)} WHERE id={esc(iid)} AND (serial IS NULL OR serial='');")
    sql_lines.append('')

    # counters UPSERT
    sql_lines.append('-- ---- rental_counters ----')
    seen = set()  # (item_id, ym) dedup
    for c in report['counters']:
        key = (c['item_id'], c['ym'])
        if key in seen: continue
        seen.add(key)
        sql_lines.append(
            f"INSERT INTO rental_counters (item_id, ym, bw, color, source) "
            f"VALUES ({esc(c['item_id'])}, {esc(c['ym'])}, {esc(c['bw'])}, {esc(c['color'])}, 'xlsx-import') "
            f"ON CONFLICT (item_id, ym) DO UPDATE SET bw=EXCLUDED.bw, color=EXCLUDED.color, source=EXCLUDED.source;"
        )
    sql_lines.append('')
    sql_lines.append(f'-- 통계: months={len(files)}, serial_updates={len(item_to_serial)}, counters={len(seen)}')

    OUT_SQL.write_text('\n'.join(sql_lines), encoding='utf-8')
    report['serial_updates'] = item_to_serial
    OUT_RPT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')

    # 콘솔 요약
    print(f'\n=== 매칭 리포트 ===')
    for ym, r in sorted(report['months'].items()):
        print(f"  {ym}: total={r['total']:3d}  matched={r['matched']:3d}  missed={r['missed']:3d}  ({r['file']})")
    print(f'\nserial UPDATE: {len(item_to_serial)}건')
    print(f'counter rows : {len(seen)}건')
    print(f'\nSQL → {OUT_SQL.relative_to(ROOT)}')
    print(f'리포트 → {OUT_RPT.relative_to(ROOT)}')

if __name__ == '__main__':
    main()
