"""
NAS 월별카운터 폴더의 한별카운터*.xlsx 모두 → 임대관리 시드 카운터 JSON.
"""
import openpyxl, json, re, zipfile, tempfile
from pathlib import Path

NAS = Path(r'\\192.168.0.251\업무공용\0 한별임대현황20260511\월별카운터')
OUT = Path(r'C:\Users\UserK\Desktop\클로드코드공부\임대관리\tools\seed_counters.json')

def to_int(v):
    if v is None or v == '' or (isinstance(v, str) and v.strip() in ('', ' ')): return None
    if isinstance(v, (int, float)): return int(v)
    s = str(v).strip().replace(',', '')
    try: return int(float(s))
    except: return None

def find_files(root):
    res = []
    if not root.exists():
        # 다른 경로 시도
        for alt in [Path(r'\\192.168.0.251\업무공용\경리업무\한별시스템 매출현황\임대현황\월별카운터')]:
            if alt.exists():
                root = alt
                break
    for p in root.rglob('한별카운터*'):
        if not p.is_file(): continue
        ext = p.suffix.lower()
        if ext == '.xlsx':
            res.append(('xlsx', p))
        elif ext == '.zip':
            res.append(('zip', p))
        elif ext == '':
            try:
                with open(p, 'rb') as f: sig = f.read(4)
                if sig.startswith(b'PK'): res.append(('zip-or-xlsx', p))
            except: pass
    return res

def parse_xlsx(xlsx):
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]

    created = None
    for r in range(1, 6):
        for c in range(1, 8):
            v = ws.cell(row=r, column=c).value
            if v and re.search(r'\d{4}[/.-]\d{1,2}[/.-]\d{1,2}', str(v)):
                created = str(v).strip(); break
        if created: break
    period = None
    if created:
        m = re.search(r'(\d{4})[/.-](\d{1,2})[/.-]', created)
        if m:
            y, mo = int(m.group(1)), int(m.group(2))
            uy = y if mo > 1 else y - 1
            um = mo - 1 if mo > 1 else 12
            period = f"{uy}-{um:02d}"

    header_row = None
    for r in range(1, min(ws.max_row + 1, 20)):
        for c in range(1, min(ws.max_column + 1, 25)):
            v = ws.cell(row=r, column=c).value
            if v and ('일련' in str(v) or 'serial' in str(v).lower()):
                header_row = r; break
        if header_row: break
    if not header_row:
        return created, period, {}, None

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v: headers[c] = str(v).replace('\n','').replace(' ','')

    cols = {}
    for c, s in headers.items():
        if s == '일련번호': cols['serial'] = c
        elif s == '모델이름': cols['model'] = c
        elif s == '그룹': cols['group'] = c
        elif s == '자산번호': cols['asset'] = c
        elif s == '흑백합계': cols['bw'] = c
        elif s == '컬러합계': cols['co'] = c
        elif s == '결합합계': cols['combined'] = c
        elif s == '마지막업데이트': cols['last_update'] = c
    EXCLUDE = ('복사기','프린터','단색','팩스')
    for key, words in [('serial',['일련']),('model',['모델']),('group',['그룹']),
                        ('asset',['자산']),('bw',['흑백']),('co',['컬러','칼라']),
                        ('last_update',['마지막업데이트','업데이트'])]:
        if key in cols: continue
        for c, s in headers.items():
            if any(w in s for w in words) and not any(x in s for x in EXCLUDE):
                cols[key] = c; break

    counters = {}
    if 'serial' not in cols:
        return created, period, counters, cols
    for r in range(header_row + 1, ws.max_row + 1):
        serial = ws.cell(row=r, column=cols['serial']).value
        if not serial: continue
        sk = str(serial).strip()
        if not sk: continue
        bw = to_int(ws.cell(row=r, column=cols['bw']).value) if cols.get('bw') else None
        co = to_int(ws.cell(row=r, column=cols['co']).value) if cols.get('co') else None
        if bw is None and co is None: continue
        counters[sk] = {
            'bw': bw, 'co': co,
            'last_update': str(ws.cell(row=r, column=cols['last_update']).value or '').strip() if cols.get('last_update') else '',
            'model': str(ws.cell(row=r, column=cols['model']).value or '').strip() if cols.get('model') else '',
            'group': str(ws.cell(row=r, column=cols['group']).value or '').strip() if cols.get('group') else '',
            'asset_name': str(ws.cell(row=r, column=cols['asset']).value or '').strip() if cols.get('asset') else '',
        }
    wb.close()
    return created, period, counters, cols

def extract_xlsx_from_zip(zp):
    with zipfile.ZipFile(zp) as z:
        names = z.namelist()
        if any(n.startswith('xl/') for n in names) or '[Content_Types].xml' in names:
            return zp
        target = next((n for n in names if '한별카운터' in n and n.lower().endswith('.xlsx')), None)
        if not target: target = next((n for n in names if n.lower().endswith('.xlsx')), None)
        if not target: return None
        tmp = Path(tempfile.gettempdir()) / f"hbs_{zp.stem}_{Path(target).name}"
        with z.open(target) as src, open(tmp, 'wb') as dst: dst.write(src.read())
        return tmp

def main():
    files = find_files(NAS)
    print(f"NAS 발견: {len(files)}")
    counters = {}
    printers = {}
    errors = []
    files_sorted = sorted(files, key=lambda x: (x[0] != 'xlsx', str(x[1])))
    for kind, path in files_sorted:
        xp = path
        if kind in ('zip', 'zip-or-xlsx'):
            try:
                xp = extract_xlsx_from_zip(path)
                if not xp: errors.append(f"{path.name}: zip 처리 실패"); continue
            except Exception as e:
                errors.append(f"{path.name}: {e}"); continue
        try:
            _, period, c, _ = parse_xlsx(xp)
        except Exception as e:
            errors.append(f"{path.name}: parse - {e}"); continue
        if not period or not c: continue
        if period in counters and len(counters[period]) >= len(c): continue
        merged = {}
        for s, info in c.items():
            merged[s] = {'bw': info['bw'], 'co': info['co'], 'last_update': info['last_update'],
                         'source': 'nas_counter_report', 'source_file': path.name}
            if s not in printers:
                printers[s] = {'serial': s, 'model': info['model'], 'group': info['group'], 'asset_name': info['asset_name']}
        counters[period] = merged

    counters_sorted = dict(sorted(counters.items()))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump({'counters': counters_sorted, 'printers': printers, 'errors': errors[:10]}, f, ensure_ascii=False, indent=2)

    keys = list(counters_sorted.keys())
    print(f"기간: {len(keys)}개월 ({keys[0] if keys else '-'} ~ {keys[-1] if keys else '-'})")
    print(f"고유 시리얼: {len(printers)}")
    print(f"오류: {len(errors)}")
    print(f"출력: {OUT}")

if __name__ == '__main__':
    main()
