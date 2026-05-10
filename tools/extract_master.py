"""
한별시스템 임대현황.xlsx (NAS 마스터) → 임대관리 시드 JSON 추출.
시트별 처리:
  - 임대현황 (275건): 메인 임대 거래처 마스터
  - 무상임대현황 (37건): 무상 임대 (참고용)
  - BNI: BNI 회원 임대 (참고용)
  - 우리정보임대: 별도 임대 (참고용)
"""
import openpyxl, json, re
from pathlib import Path

SRC = Path(r'C:\Users\UserK\Desktop\클로드코드공부\_temp_inspect\rental_master_open.xlsx')
OUT = Path(r'C:\Users\UserK\Desktop\클로드코드공부\임대관리\tools\seed_customers.json')

def to_int(v):
    if v is None or v == '': return None
    if isinstance(v, (int, float)): return int(v)
    s = str(v).strip().replace(',', '').replace(' ', '')
    if not s: return None
    try: return int(float(s))
    except: return None

def clean(v):
    if v is None: return ''
    return str(v).strip()

def parse_extra_charge(text):
    """추가장당 필드 파싱 → bw_free/bw_rate/co_free/co_rate
    예시:
      '흑 무제한\n컬러구매'
      '흑:3,000매(10원)'
      '흑:1,000매(10원)\n컬러:500매(100원)'
      '흑:5,000매(10원)\n컬러:1,000매(100원)'
    """
    if not text: return {}
    t = str(text).replace(' ', '').replace(',', '')
    out = {}
    # 흑백
    m = re.search(r'흑[:\s]*(\d+)매?\((\d+)원\)', t)
    if m:
        out['bw_free'] = int(m.group(1))
        out['bw_rate'] = int(m.group(2))
    elif '흑무제한' in t or '흑백무제한' in t:
        out['bw_free'] = 999999
        out['bw_rate'] = 0
    # 컬러
    m = re.search(r'(?:컬러|칼라|색상)[:\s]*(\d+)매?\((\d+)원\)', t)
    if m:
        out['co_free'] = int(m.group(1))
        out['co_rate'] = int(m.group(2))
    elif '컬러구매' in t or '컬러구매조건' in t:
        out['co_free'] = 0
        out['co_rate'] = 0  # 별도
        out['co_note'] = '컬러구매'
    return out

def parse_billing_day(text):
    """청구일 / 입금일 추출"""
    if not text: return None
    m = re.search(r'(\d{1,2})', str(text))
    return int(m.group(1)) if m else None

def extract_main_sheet(ws):
    """임대현황 시트 — R6 헤더, R7~ 데이터"""
    rows = []
    for r in range(7, ws.max_row + 1):
        no = ws.cell(row=r, column=1).value
        company = clean(ws.cell(row=r, column=2).value)
        if not company: continue
        # 합계 행 등 제외
        if '합계' in company or '계' == company.strip(): continue

        device     = clean(ws.cell(row=r, column=3).value)
        amount     = to_int(ws.cell(row=r, column=4).value)        # 임대금액
        amount_vat = to_int(ws.cell(row=r, column=5).value)        # 사업자금액(VAT포함)
        check_day  = clean(ws.cell(row=r, column=6).value)         # 카운터점검일
        bill_day   = clean(ws.cell(row=r, column=7).value)         # 청구일
        invoice    = clean(ws.cell(row=r, column=8).value)         # 계산서
        deposit_day= clean(ws.cell(row=r, column=9).value)         # 입금일(익월)
        extra_text = clean(ws.cell(row=r, column=10).value)        # 추가장당
        address    = clean(ws.cell(row=r, column=11).value)
        rent_date  = clean(ws.cell(row=r, column=12).value)        # 임대날짜 (시작일\n갱신일)
        composition= clean(ws.cell(row=r, column=13).value)        # 제품구성
        prepay     = clean(ws.cell(row=r, column=14).value)        # 선후불/보증금
        memo       = clean(ws.cell(row=r, column=15).value)

        rates = parse_extra_charge(extra_text)

        rows.append({
            'no': to_int(no),
            'company': company,
            'device': device,
            'base_fee': amount,
            'base_fee_vat': amount_vat,
            'check_day': check_day,
            'bill_day_text': bill_day,
            'bill_day': parse_billing_day(bill_day),
            'invoice_issued': invoice,
            'deposit_day': deposit_day,
            'extra_charge_text': extra_text,
            **rates,
            'address': address,
            'rent_date': rent_date,
            'composition': composition,
            'prepay_deposit': prepay,
            'memo': memo,
            'source_sheet': '임대현황',
        })
    return rows

def extract_free_sheet(ws):
    """무상임대현황 — R3 ~"""
    rows = []
    for r in range(3, ws.max_row + 1):
        company = clean(ws.cell(row=r, column=2).value)
        device  = clean(ws.cell(row=r, column=3).value)
        if not (company or device): continue
        rows.append({
            'company': company,
            'device': device,
            'base_fee': 0,
            'rent_date': clean(ws.cell(row=r, column=5).value),
            'memo': '무상 임대 · ' + clean(ws.cell(row=r, column=6).value),
            'source_sheet': '무상임대현황',
        })
    return rows

def extract_bni_sheet(ws):
    """BNI — R5 ~"""
    rows = []
    for r in range(5, ws.max_row + 1):
        company = clean(ws.cell(row=r, column=2).value)
        device  = clean(ws.cell(row=r, column=3).value)
        if not (company or device): continue
        rows.append({
            'company': company,
            'device': device,
            'base_fee': to_int(ws.cell(row=r, column=4).value),
            'memo': 'BNI · ' + clean(ws.cell(row=r, column=5).value),
            'rent_date': clean(ws.cell(row=r, column=7).value),
            'source_sheet': 'BNI',
        })
    return rows

def extract_woori_sheet(ws):
    """우리정보임대 — R3 ~"""
    rows = []
    for r in range(3, ws.max_row + 1):
        company = clean(ws.cell(row=r, column=2).value)
        if not company: continue
        rows.append({
            'company': company,
            'device': clean(ws.cell(row=r, column=3).value),
            'rent_date': clean(ws.cell(row=r, column=4).value),
            'base_fee': to_int(ws.cell(row=r, column=7).value),
            'phone': clean(ws.cell(row=r, column=9).value),
            'address': clean(ws.cell(row=r, column=10).value),
            'composition': clean(ws.cell(row=r, column=11).value),
            'memo': '우리정보 · ' + clean(ws.cell(row=r, column=14).value or ''),
            'source_sheet': '우리정보임대',
        })
    return rows

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    main_rows  = extract_main_sheet(wb['임대현황'])
    free_rows  = extract_free_sheet(wb['무상임대현황'])
    bni_rows   = extract_bni_sheet(wb['BNI'])
    woori_rows = extract_woori_sheet(wb['우리정보임대'])

    all_rows = main_rows + free_rows + bni_rows + woori_rows

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump({
            'extracted_at': '2026-05-11',
            'source_file': '한별시스템 임대현황.xlsx',
            'counts': {
                '임대현황': len(main_rows),
                '무상임대현황': len(free_rows),
                'BNI': len(bni_rows),
                '우리정보임대': len(woori_rows),
                'total': len(all_rows),
            },
            'rows': all_rows,
        }, f, ensure_ascii=False, indent=2)

    print(f"=== 추출 완료 ===")
    print(f"  임대현황:     {len(main_rows)}건")
    print(f"  무상임대현황: {len(free_rows)}건")
    print(f"  BNI:          {len(bni_rows)}건")
    print(f"  우리정보임대: {len(woori_rows)}건")
    print(f"  합계:         {len(all_rows)}건")
    print(f"\n출력: {OUT}")

    # 샘플
    print("\n=== 임대현황 샘플 5건 ===")
    for r in main_rows[:5]:
        print(f"  {r['no']:3} {r['company']:20} | 임대료 {r.get('base_fee')} | 흑백 {r.get('bw_free')}매/{r.get('bw_rate')}원 | 컬러 {r.get('co_free')}매/{r.get('co_rate')}원")

if __name__ == '__main__':
    main()
